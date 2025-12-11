#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
阿里国际业务智能复盘工具
Ali International Business Intelligence Analysis Tool

功能：
1. Excel数据读取和分析
2. AI智能分类（通义千问API）
3. 数据可视化
4. 报告生成
5. 智能提醒
6. GUI界面
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
# seaborn 作为可选依赖
try:
    import seaborn as sns
    SEABORN_AVAILABLE = True
except ImportError:
    SEABORN_AVAILABLE = False
import json
import logging
import argparse
import os
import sys
from datetime import datetime, timedelta
from typing import Dict, List, Tuple, Optional
import requests
import warnings
from pathlib import Path
# tkinter 作为可选依赖（网页版不需要）
try:
    import tkinter as tk
    from tkinter import ttk, filedialog, messagebox, scrolledtext
    TKINTER_AVAILABLE = True
except ImportError:
    TKINTER_AVAILABLE = False
    # 创建占位符以避免错误
    tk = None
    ttk = None
    filedialog = None
    messagebox = None
    scrolledtext = None
import threading
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from docx import Document
from docx.shared import Inches
import plotly.graph_objects as go
import plotly.express as px
from plotly.offline import plot
import folium
from folium import plugins

# === 中文字体支持 - 使用FontProperties直接指定字体文件 ===
import matplotlib
matplotlib.use('Agg')  # 使用非交互式后端
import matplotlib.pyplot as plt
from matplotlib import rcParams
import matplotlib.font_manager as fm
import os

# 全局字体配置 - 直接使用字体文件路径
CHINESE_FONT_PATH = r'C:\Windows\Fonts\simhei.ttf'  # 黑体
CHINESE_FONT_PROP = None

if os.path.exists(CHINESE_FONT_PATH):
    try:
        CHINESE_FONT_PROP = fm.FontProperties(fname=CHINESE_FONT_PATH)
        print(f"[OK] 加载中文字体: {CHINESE_FONT_PROP.get_name()} from {CHINESE_FONT_PATH}")
    except Exception as e:
        print(f"[WARNING] 加载字体失败: {e}")
        CHINESE_FONT_PROP = None

# 同时设置rcParams作为备用
plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'SimSun']
plt.rcParams['axes.unicode_minus'] = False
plt.rcParams['font.family'] = 'sans-serif'

# =============================================

# 忽略警告
warnings.filterwarnings('ignore')

# 辅助函数：为matplotlib图表设置中文字体
def set_chinese_font_for_plot(ax, title=None, xlabel=None, ylabel=None):
    """为matplotlib图表的标题、标签设置中文字体"""
    if CHINESE_FONT_PROP:
        if title:
            ax.set_title(title, fontproperties=CHINESE_FONT_PROP, fontsize=16, fontweight='bold', pad=20, color='black')
        if xlabel:
            ax.set_xlabel(xlabel, fontproperties=CHINESE_FONT_PROP, fontsize=12, color='black')
        if ylabel:
            ax.set_ylabel(ylabel, fontproperties=CHINESE_FONT_PROP, fontsize=12, color='black')
        
        # 设置刻度标签字体
        for label in ax.get_xticklabels():
            label.set_fontproperties(CHINESE_FONT_PROP)
            label.set_color('black')
        for label in ax.get_yticklabels():
            label.set_fontproperties(CHINESE_FONT_PROP)
            label.set_color('black')
    else:
        if title:
            ax.set_title(title, fontsize=16, fontweight='bold', pad=20, color='black')
        if xlabel:
            ax.set_xlabel(xlabel, fontsize=12, color='black')
        if ylabel:
            ax.set_ylabel(ylabel, fontsize=12, color='black')

class AliBusinessAnalyzer:
    """阿里国际业务智能复盘工具主类"""
    
    def __init__(self, config_file: str = "config.json"):
        """初始化分析器"""
        self.config = self._load_config(config_file)
        self.data = None
        self.analysis_results = {}
        self.setup_logging()
        
        # 标准列名（完全匹配Excel模板）
        self.standard_columns = [
            '询盘时间', '咨询方式', '跟进等级', '客户名称', '客户层级', 
            '所属大洲', '国家', '询价产品', '产品ID', '跟进人', 
            '备注 (失单原因+跟进机会点)', '最后跟进时间'
        ]
        
        # 初始化图表样式
        self._setup_plot_style()
        
    def _load_config(self, config_file: str) -> Dict:
        """加载配置文件"""
        default_config = {
            "api_key": "sk-9c3866e9c45b4e5ea89faa1796fe78ff",
            "api_url": "https://dashscope.aliyuncs.com/api/v1/services/aigc/text-generation/generation",
            "data_path": "./data",
            "output_path": "./output",
            "log_level": "INFO",
            "max_retries": 3,
            "timeout": 30,
            "ai_model": "qwen-turbo",
            "classification_rules": {
                "A": "精准询盘：客户明确指出产品需求，包含各种信息（数量、运输/支付要求、公司信息等）",
                "B": "普通询盘：广撒网询盘，内容广泛，只是询价或发对产品感兴趣，或信息未读，需要继续跟进了",
                "C": "个人买家/不匹配询盘/垃圾询盘",
                "X": "已下样品单/大货客户，持续跟进"
            }
        }
        
        if os.path.exists(config_file):
            try:
                with open(config_file, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                # 合并默认配置
                for key, value in default_config.items():
                    if key not in config:
                        config[key] = value
                return config
            except Exception as e:
                print(f"配置文件加载失败，使用默认配置: {e}")
                return default_config
        else:
            # 创建默认配置文件
            with open(config_file, 'w', encoding='utf-8') as f:
                json.dump(default_config, f, ensure_ascii=False, indent=2)
            return default_config
    
    def setup_logging(self):
        """设置日志"""
        log_level = getattr(logging, self.config.get('log_level', 'INFO'))
        
        # 设置UTF-8编码的StreamHandler
        import sys
        stream_handler = logging.StreamHandler(sys.stdout)
        stream_handler.setFormatter(logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s'))
        
        logging.basicConfig(
            level=log_level,
            format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler('ali_business_analyzer.log', encoding='utf-8'),
                stream_handler
            ]
        )
        self.logger = logging.getLogger(__name__)
        self.logger.info("阿里国际业务智能复盘工具启动")
    
    def _setup_plot_style(self):
        """设置图表样式"""
        try:
            if SEABORN_AVAILABLE:
                sns.set_style("whitegrid")
            # 尝试使用 seaborn 样式，如果失败则使用默认样式
            try:
                plt.style.use('seaborn-v0_8')
            except (OSError, KeyError):
                try:
                    plt.style.use('seaborn')
                except (OSError, KeyError):
                    # 如果都不行，使用默认样式
                    pass
        except Exception as e:
            # 如果设置样式失败，使用默认样式，不影响程序运行
            pass
        
    def read_excel(self, file_path: str, sheet_name: str = None) -> pd.DataFrame:
        """读取Excel文件 - 支持多个工作表"""
        try:
            self.logger.info(f"正在读取Excel文件: {file_path}")
            
            # 尝试读取Excel文件
            if file_path.endswith('.xlsx') or file_path.endswith('.xls'):
                # 读取所有工作表
                excel_file = pd.ExcelFile(file_path)
                sheet_names = excel_file.sheet_names
                self.logger.info(f"发现 {len(sheet_names)} 个工作表: {sheet_names}")
                
                # 如果指定了工作表，只读取指定的
                if sheet_name:
                    if sheet_name not in sheet_names:
                        raise ValueError(f"工作表 '{sheet_name}' 不存在")
                    df = pd.read_excel(file_path, sheet_name=sheet_name)
                    self.logger.info(f"读取工作表: {sheet_name}")
                else:
                    # 读取所有工作表并合并
                    all_dfs = []
                    for sheet in sheet_names:
                        try:
                            df_sheet = pd.read_excel(file_path, sheet_name=sheet)
                            # 添加工作表名称列（可选）
                            df_sheet['_来源工作表'] = sheet
                            all_dfs.append(df_sheet)
                            self.logger.info(f"从工作表 '{sheet}' 读取了 {len(df_sheet)} 条记录")
                        except Exception as e:
                            self.logger.error(f"读取工作表 '{sheet}' 失败: {e}")
                            continue
                    
                    if not all_dfs:
                        raise ValueError("没有成功读取任何工作表")
                    
                    # 合并所有数据
                    df = pd.concat(all_dfs, ignore_index=True)
                    self.logger.info(f"合并后共 {len(df)} 条记录")
                    
            else:
                raise ValueError("文件格式不支持，请使用Excel文件(.xlsx或.xls)")
            
            # 标准化列名
            df = self._standardize_columns(df)
            
            # 数据清洗
            df = self._clean_data(df)
            
            # 转换日期格式（去掉时间部分）
            df = self._convert_date_format(df)
            
            self.data = df
            self.logger.info(f"成功读取 {len(df)} 条记录")
            return df
            
        except Exception as e:
            self.logger.error(f"读取Excel文件失败: {e}")
            raise
    
    def _convert_date_format(self, df: pd.DataFrame) -> pd.DataFrame:
        """转换日期格式，只保留日期，不包含时间"""
        date_columns = ['询盘时间', '最后跟进时间']
        
        for col in date_columns:
            if col in df.columns:
                # 尝试转换日期格式
                try:
                    # 先转换为datetime
                    df[col] = pd.to_datetime(df[col], errors='coerce')
                    # 只对有效的datetime进行格式化
                    mask = df[col].notna()
                    if mask.any():
                        df.loc[mask, col] = df.loc[mask, col].dt.strftime('%Y/%m/%d')
                    # 保持无效日期为空字符串而不是NaN
                    df[col] = df[col].fillna('')
                except Exception as e:
                    self.logger.warning(f"日期列 '{col}' 转换失败: {e}")
        
        return df
    
    def _standardize_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """标准化列名"""
        column_mapping = {}
        used_standard_columns = set()
        
        # 创建列名映射
        for col in df.columns:
            col_lower = str(col).lower().strip()
            
            # 匹配标准列名
            if any(keyword in col_lower for keyword in ['询盘', '时间', 'inquiry', 'time']) and '询盘时间' not in used_standard_columns:
                column_mapping[col] = '询盘时间'
                used_standard_columns.add('询盘时间')
            elif any(keyword in col_lower for keyword in ['咨询', '方式', 'contact', 'method']) and '咨询方式' not in used_standard_columns:
                column_mapping[col] = '咨询方式'
                used_standard_columns.add('咨询方式')
            elif any(keyword in col_lower for keyword in ['跟进', '等级', 'level', 'grade']) and '跟进等级' not in used_standard_columns:
                column_mapping[col] = '跟进等级'
                used_standard_columns.add('跟进等级')
            elif any(keyword in col_lower for keyword in ['客户', '名称', 'customer', 'name']) and '客户名称' not in used_standard_columns:
                column_mapping[col] = '客户名称'
                used_standard_columns.add('客户名称')
            elif any(keyword in col_lower for keyword in ['层级', 'tier', 'category']) and '客户层级' not in used_standard_columns:
                column_mapping[col] = '客户层级'
                used_standard_columns.add('客户层级')
            elif any(keyword in col_lower for keyword in ['大洲', 'continent']) and '所属大洲' not in used_standard_columns:
                column_mapping[col] = '所属大洲'
                used_standard_columns.add('所属大洲')
            elif any(keyword in col_lower for keyword in ['国家', 'country', 'nation']) and '国家' not in used_standard_columns:
                column_mapping[col] = '国家'
                used_standard_columns.add('国家')
            elif any(keyword in col_lower for keyword in ['产品', 'product', '询价']) and '询价产品' not in used_standard_columns:
                column_mapping[col] = '询价产品'
                used_standard_columns.add('询价产品')
            elif any(keyword in col_lower for keyword in ['id', '产品id']) and '产品ID' not in used_standard_columns:
                column_mapping[col] = '产品ID'
                used_standard_columns.add('产品ID')
            elif any(keyword in col_lower for keyword in ['跟进人', 'follower', 'handler']) and '跟进人' not in used_standard_columns:
                column_mapping[col] = '跟进人'
                used_standard_columns.add('跟进人')
            elif any(keyword in col_lower for keyword in ['备注', 'remark', 'note']) and '备注 (失单原因+跟进机会点)' not in used_standard_columns:
                column_mapping[col] = '备注 (失单原因+跟进机会点)'
                used_standard_columns.add('备注 (失单原因+跟进机会点)')
            elif any(keyword in col_lower for keyword in ['最后', 'last', 'follow']) and '最后跟进时间' not in used_standard_columns:
                column_mapping[col] = '最后跟进时间'
                used_standard_columns.add('最后跟进时间')
        
        # 重命名列
        df = df.rename(columns=column_mapping)
        
        # 添加缺失的列
        for col in self.standard_columns:
            if col not in df.columns:
                df[col] = None
        
        return df
    
    def _clean_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """数据清洗"""
        original_count = len(df)
        self.logger.info(f"清洗前数据量: {original_count}")
        
        # 处理时间列
        time_columns = ['询盘时间', '最后跟进时间']
        for col in time_columns:
            if col in df.columns:
                try:
                    # 检查是否是Excel日期序列号（数值类型）
                    if df[col].dtype in ['float64', 'int64']:
                        # Excel日期从1900-01-01开始，序列号0对应1900-01-01
                        # 需要减去60以修正Excel的日期起始错误（Excel认为1900是闰年）
                        dates = df[col].dropna()
                        if len(dates) > 0:
                            # 尝试转换为日期
                            try:
                                # 如果是合理的日期序列号（大于25569表示Excel日期，小于25569可能是文本）
                                if dates.min() > 25568:
                                    # Excel日期序列号
                                    df[col] = df[col].apply(lambda x: pd.to_datetime('1899-12-30') + pd.Timedelta(days=x) if pd.notna(x) and x > 0 else None)
                                else:
                                    # 常规日期转换
                                    df[col] = pd.to_datetime(df[col], errors='coerce')
                            except:
                                df[col] = pd.to_datetime(df[col], errors='coerce')
                        else:
                            df[col] = pd.to_datetime(df[col], errors='coerce')
                    else:
                        # 其他类型，正常转换
                        df[col] = pd.to_datetime(df[col], errors='coerce')
                except Exception as e:
                    self.logger.warning(f"处理时间列 '{col}' 时出错: {e}")
        
        # 只替换数字类型为NaN的空值
        df = df.fillna('')
        
        # 去除完全重复的行（基于所有列）
        df = df.drop_duplicates()
        
        final_count = len(df)
        self.logger.info(f"清洗后数据量: {final_count}")
        
        return df
    
    def ai_classify_customer(self, customer_data: Dict) -> Dict:
        """使用AI对客户进行分类"""
        try:
            # 构建提示词
            prompt = self._build_classification_prompt(customer_data)
            
            # 调用通义千问API
            response = self._call_qwen_api(prompt)
            
            # 解析响应
            result = self._parse_classification_response(response)
            
            return result
            
        except Exception as e:
            self.logger.error(f"AI分类失败: {e}")
            return {
                'classification': 'C',
                'intent': '未知',
                'suggestion': '需要进一步了解客户需求'
            }
    
    def _build_classification_prompt(self, customer_data: Dict) -> str:
        """构建分类提示词"""
        prompt = f"""
请根据以下客户信息进行智能分类：

客户信息：
- 客户名称：{customer_data.get('客户名称', '未知')}
- 咨询方式：{customer_data.get('咨询方式', '未知')}
- 询价产品：{customer_data.get('询价产品', '未知')}
- 所属大洲：{customer_data.get('所属大洲', '未知')}
- 国家：{customer_data.get('国家', '未知')}
- 备注：{customer_data.get('备注 (失单原因+跟进机会点)', '无')}

请按照以下规则进行分类：

A类：精准询盘 - 客户明确指出产品需求，包含各种信息（数量、运输/支付要求、公司信息等）
B类：普通询盘 - 广撒网询盘，内容广泛，只是询价或发对产品感兴趣，或信息未读，需要继续跟进了
C类：个人买家/不匹配询盘/垃圾询盘
X类：已下样品单/大货客户，持续跟进

同时请分析：
1. 客户意图（如：采购、样品、价格咨询、技术咨询等）
2. 跟进建议（具体的行动建议）

请以JSON格式返回结果：
{{
    "classification": "A/B/C/X",
    "intent": "客户意图",
    "suggestion": "跟进建议"
}}
"""
        return prompt
    
    def _call_qwen_api(self, prompt: str) -> str:
        """调用通义千问API"""
        headers = {
            'Authorization': f'Bearer {self.config["api_key"]}',
            'Content-Type': 'application/json'
        }
        
        data = {
            'model': self.config['ai_model'],
            'input': {
                'messages': [
                    {
                        'role': 'user',
                        'content': prompt
                    }
                ]
            },
            'parameters': {
                'temperature': 0.7,
                'max_tokens': 1000
            }
        }
        
        try:
            response = requests.post(
                self.config['api_url'],
                headers=headers,
                json=data,
                timeout=self.config['timeout']
            )
            
            if response.status_code == 200:
                result = response.json()
                return result['output']['text']
            else:
                self.logger.error(f"API调用失败: {response.status_code}")
                return ""
                
        except Exception as e:
            self.logger.error(f"API调用异常: {e}")
            return ""
    
    def _parse_classification_response(self, response: str) -> Dict:
        """解析AI响应"""
        try:
            # 尝试解析JSON
            import re
            json_match = re.search(r'\{.*\}', response, re.DOTALL)
            if json_match:
                json_str = json_match.group()
                return json.loads(json_str)
            else:
                # 如果无法解析JSON，使用正则表达式提取信息
                classification = 'C'
                intent = '未知'
                suggestion = '需要进一步了解客户需求'
                
                if 'A' in response:
                    classification = 'A'
                elif 'B' in response:
                    classification = 'B'
                elif 'X' in response:
                    classification = 'X'
                
                return {
                    'classification': classification,
                    'intent': intent,
                    'suggestion': suggestion
                }
                
        except Exception as e:
            self.logger.error(f"解析AI响应失败: {e}")
            return {
                'classification': 'C',
                'intent': '未知',
                'suggestion': '需要进一步了解客户需求'
            }
    
    def analyze_data(self) -> Dict:
        """数据分析"""
        if self.data is None:
            raise ValueError("请先读取数据")
        
        self.logger.info("开始数据分析")
        
        analysis_results = {}
        
        # 1. 基本统计
        analysis_results['basic_stats'] = {
            'total_customers': len(self.data),
            'total_inquiries': len(self.data),
            'date_range': {
                'start': self.data['询盘时间'].min() if '询盘时间' in self.data.columns else None,
                'end': self.data['询盘时间'].max() if '询盘时间' in self.data.columns else None
            }
        }
        
        # 2. 地区分析
        if '所属大洲' in self.data.columns:
            analysis_results['continent_analysis'] = self.data['所属大洲'].value_counts().to_dict()
        
        if '国家' in self.data.columns:
            analysis_results['country_analysis'] = self.data['国家'].value_counts().head(10).to_dict()
        
        # 3. 产品分析
        if '询价产品' in self.data.columns:
            analysis_results['product_analysis'] = self.data['询价产品'].value_counts().head(10).to_dict()
        
        # 4. 跟进等级分析
        if '跟进等级' in self.data.columns:
            analysis_results['follow_up_analysis'] = self.data['跟进等级'].value_counts().to_dict()
        
        # 5. 时间趋势分析
        if '询盘时间' in self.data.columns:
            try:
                # 确保时间是datetime类型
                time_series = pd.to_datetime(self.data['询盘时间'], errors='coerce')
                # 移除无效日期
                valid_times = time_series.dropna()
                if len(valid_times) > 0:
                    daily_inquiries = valid_times.groupby(valid_times.dt.date).size()
                    analysis_results['daily_trend'] = daily_inquiries.to_dict()
                else:
                    analysis_results['daily_trend'] = {}
            except Exception as e:
                self.logger.warning(f"时间趋势分析失败: {e}")
                analysis_results['daily_trend'] = {}
        
        # 6. 跟进人分析
        if '跟进人' in self.data.columns:
            analysis_results['handler_analysis'] = self.data['跟进人'].value_counts().to_dict()
        
        self.analysis_results = analysis_results
        self.logger.info("数据分析完成")
        
        return analysis_results
    
    def generate_visualizations(self, output_dir: str = "./output"):
        """生成可视化图表"""
        if self.data is None:
            raise ValueError("请先读取数据")
        
        os.makedirs(output_dir, exist_ok=True)
        
        # 1. 地区分布图 (Continent Distribution)
        if '所属大洲' in self.data.columns:
            fig, ax = plt.subplots(figsize=(12, 8))
            continent_counts = self.data['所属大洲'].value_counts()
            
            colors = ['#FF6B6B', '#4ECDC4', '#45B7D1', '#96CEB4', '#FFEAA7', '#DDA0DD']
            wedges, texts, autotexts = ax.pie(
                continent_counts.values, 
                labels=continent_counts.index, 
                autopct='%1.1f%%',
                colors=colors, 
                startangle=90,
                textprops={'fontsize': 12, 'color': 'black'}
            )
            
            # 设置中文字体和颜色
            for text in texts:
                if CHINESE_FONT_PROP:
                    text.set_fontproperties(CHINESE_FONT_PROP)
                text.set_color('black')
                text.set_fontsize(14)
                text.set_fontweight('bold')
            for autotext in autotexts:
                autotext.set_color('white')
                autotext.set_fontweight('bold')
                autotext.set_fontsize(10)
            
            # 设置标题
            if CHINESE_FONT_PROP:
                ax.set_title('客户地区分布 (Customer Regional Distribution)', 
                           fontproperties=CHINESE_FONT_PROP, fontsize=16, fontweight='bold', pad=20, color='black')
            else:
                ax.set_title('客户地区分布 (Customer Regional Distribution)', 
                           fontsize=16, fontweight='bold', pad=20, color='black')
            
            plt.tight_layout()
            plt.savefig(f"{output_dir}/continent_distribution.png", dpi=300, bbox_inches='tight', pad_inches=0.2)
            plt.close()
            print(f"[OK] 地区分布图已生成")
        
        # 2. 国家分布图 (Country Distribution)
        if '国家' in self.data.columns:
            fig, ax = plt.subplots(figsize=(15, 8))
            country_counts = self.data['国家'].value_counts().head(15)
            
            bars = ax.bar(range(len(country_counts)), country_counts.values, color='steelblue', alpha=0.8)
            ax.set_xticks(range(len(country_counts)))
            ax.set_xticklabels(country_counts.index, rotation=45, ha='right', fontsize=11)
            
            # 使用辅助函数设置中文字体
            set_chinese_font_for_plot(ax, 
                                    title='Top 15 国家客户分布 (Top 15 Countries)',
                                    xlabel='国家 (Country)',
                                    ylabel='客户数量 (Customer Count)')
            
            ax.grid(axis='y', alpha=0.3, linestyle='--')
            ax.tick_params(axis='y', labelcolor='black')
            
            # 在柱子上添加数值
            for i, bar in enumerate(bars):
                height = bar.get_height()
                ax.text(bar.get_x() + bar.get_width()/2., height,
                       f'{int(height)}',
                       ha='center', va='bottom', fontsize=9, color='black')
            
            plt.tight_layout()
            plt.savefig(f"{output_dir}/country_distribution.png", dpi=300, bbox_inches='tight', pad_inches=0.2)
            plt.close()
            print(f"[OK] 国家分布图已生成")
        
        # 3. 产品热度图 (Product Popularity)
        if '询价产品' in self.data.columns:
            fig, ax = plt.subplots(figsize=(15, 8))
            product_counts = self.data['询价产品'].value_counts().head(15)
            
            bars = ax.barh(range(len(product_counts)), product_counts.values, color='coral', alpha=0.8)
            ax.set_yticks(range(len(product_counts)))
            ax.set_yticklabels(product_counts.index, fontsize=11)
            
            # 使用辅助函数设置中文字体
            set_chinese_font_for_plot(ax,
                                    title='Top 15 热门产品 (Top 15 Popular Products)',
                                    xlabel='询盘次数 (Inquiry Count)',
                                    ylabel=None)
            
            ax.grid(axis='x', alpha=0.3, linestyle='--')
            ax.tick_params(axis='x', labelcolor='black')
            ax.invert_yaxis()  # 让最大值显示在最上面
            
            # 在柱子上添加数值
            for i, bar in enumerate(bars):
                width = bar.get_width()
                ax.text(width, bar.get_y() + bar.get_height()/2.,
                       f' {int(width)}',
                       ha='left', va='center', fontsize=9, color='black')
            
            plt.tight_layout()
            plt.savefig(f"{output_dir}/product_popularity.png", dpi=300, bbox_inches='tight', pad_inches=0.2)
            plt.close()
            print(f"[OK] 产品热度图已生成")
        
        # 4. 时间趋势图 (Daily Inquiry Trend)
        if '询盘时间' in self.data.columns:
            fig, ax = plt.subplots(figsize=(15, 8))
            daily_inquiries = self.data.groupby(self.data['询盘时间'].dt.date).size()
            
            line = ax.plot(daily_inquiries.index, daily_inquiries.values, 
                          marker='o', linewidth=2, color='green', alpha=0.7, label='每日询盘')[0]
            ax.fill_between(daily_inquiries.index, daily_inquiries.values, alpha=0.2, color='green')
            
            # 使用辅助函数设置中文字体
            set_chinese_font_for_plot(ax,
                                    title='每日询盘趋势 (Daily Inquiry Trend)',
                                    xlabel='日期 (Date)',
                                    ylabel='询盘数量 (Inquiry Count)')
            
            ax.tick_params(axis='x', rotation=45, labelcolor='black')
            ax.tick_params(axis='y', labelcolor='black')
            ax.grid(True, alpha=0.3, linestyle='--')
            
            # 设置图例字体
            if CHINESE_FONT_PROP:
                ax.legend(prop=CHINESE_FONT_PROP)
            else:
                ax.legend(prop={'size': 10})
            
            plt.tight_layout()
            plt.savefig(f"{output_dir}/daily_trend.png", dpi=300, bbox_inches='tight', pad_inches=0.2)
            plt.close()
            print(f"[OK] 时间趋势图已生成")
        
        # 5. 跟进等级分布 (Follow-up Level Distribution)
        if '跟进等级' in self.data.columns:
            fig, ax = plt.subplots(figsize=(10, 8))
            follow_up_counts = self.data['跟进等级'].value_counts()
            
            colors = ['#ff9999', '#66b3ff', '#99ff99', '#ffcc99']
            wedges, texts, autotexts = ax.pie(
                follow_up_counts.values, 
                labels=follow_up_counts.index, 
                autopct='%1.1f%%',
                colors=colors, 
                startangle=90,
                textprops={'fontsize': 12, 'color': 'black'}
            )
            
            # 设置中文字体和颜色
            for text in texts:
                if CHINESE_FONT_PROP:
                    text.set_fontproperties(CHINESE_FONT_PROP)
                text.set_color('black')
                text.set_fontsize(14)
                text.set_fontweight('bold')
            for autotext in autotexts:
                autotext.set_color('white')
                autotext.set_fontweight('bold')
                autotext.set_fontsize(10)
            
            # 设置标题
            if CHINESE_FONT_PROP:
                ax.set_title('跟进等级分布 (Follow-up Level Distribution)', 
                           fontproperties=CHINESE_FONT_PROP, fontsize=16, fontweight='bold', pad=20, color='black')
            else:
                ax.set_title('跟进等级分布 (Follow-up Level Distribution)', 
                           fontsize=16, fontweight='bold', pad=20, color='black')
            
            plt.tight_layout()
            plt.savefig(f"{output_dir}/follow_up_distribution.png", dpi=300, bbox_inches='tight', pad_inches=0.2)
            plt.close()
            print(f"[OK] 跟进等级分布图已生成")
        
        self.logger.info(f"可视化图表已保存到: {output_dir}")
    
    def generate_report(self, output_file: str = None, force_reanalyze: bool = True, user_date_range: tuple = None) -> str:
        """生成AI智能分析报告（纯文本格式）
        
        Args:
            output_file: 输出文件路径
            force_reanalyze: 是否强制重新分析数据（默认True，确保使用当前数据）
            user_date_range: 用户选择的时间范围 (start_date, end_date)，用于在报告中显示
        """
        # 强制重新分析当前数据，确保使用过滤后的数据
        if force_reanalyze or not self.analysis_results:
            self.analyze_data()
        
        # 只生成TXT格式报告
        if output_file is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            text_file = f"./output/analysis_report_{timestamp}.txt"
        else:
            # 确保是.txt文件
            text_file = output_file.replace('.pdf', '.txt')
        
        # 确保输出目录存在
        output_dir = os.path.dirname(text_file)
        if output_dir:
            os.makedirs(output_dir, exist_ok=True)
        
        # 直接生成AI智能分析文本报告
        self._generate_text_report(text_file, user_date_range)
        
        self.logger.info(f"AI智能分析报告已生成: {text_file}")
        
        return text_file  # 返回文本文件路径供GUI使用
    
    def _call_qianwen_api(self, data_summary: str) -> str:
        """调用阿里千问API进行智能分析"""
        try:
            # 阿里千问API配置
            api_key = os.getenv('DASHSCOPE_API_KEY', 'sk-09641de5f87c432b8f81c115bb0ab18a')  # 有效的API Key
            api_url = 'https://dashscope.aliyuncs.com/api/v1/services/aigc/text-generation/generation'
            
            # 构建提示词
            prompt = f"""你是一位资深的阿里国际站数据分析专家，拥有10年以上的跨境电商数据分析经验。
请基于以下询盘数据，提供一份专业、详细、有洞察力的业务分析报告。

【数据概况】
{data_summary}

【分析要求】
1. 深度分析市场趋势和客户行为模式
2. 识别业务中的关键问题和风险点
3. 提供具体可执行的优化建议
4. 对比历史数据（如果有周期性数据），分析增长或下滑原因
5. 预测未来趋势，提出战略规划建议

【报告结构】
请按照以下结构输出，内容要专业、具体、有数据支撑：

一、执行摘要（核心指标、整体评价、关键发现）
二、市场分析（地域分布、市场机会、竞争态势）
三、产品分析（热门产品、产品组合、优化方向）
四、客户质量分析（层级分布、转化率、客户价值）
五、时间趋势分析（询盘趋势、周期性特征、增长驱动因素）
六、团队绩效分析（成员表现、协作效率、培训需求）
七、问题诊断与风险预警（现存问题、潜在风险、应对方案）
八、战略行动建议（短期1-2周、中期1-3月、长期3-6月）
九、总结与展望（整体评价、关键成功因素、未来方向）

注意：
1. 所有分析必须基于提供的真实数据
2. 每个结论都要有数据支撑
3. 建议要具体可执行，不要泛泛而谈
4. 使用专业的跨境电商术语
5. 保持客观、理性的分析态度"""

            # API请求
            headers = {
                'Authorization': f'Bearer {api_key}',
                'Content-Type': 'application/json'
            }
            
            payload = {
                "model": "qwen-max",  # 使用千问最强模型
                "input": {
                    "messages": [
                        {
                            "role": "system",
                            "content": "你是一位资深的阿里国际站数据分析专家，擅长从数据中挖掘商业洞察，提供专业的业务建议。"
                        },
                        {
                            "role": "user",
                            "content": prompt
                        }
                    ]
                },
                "parameters": {
                    "result_format": "message",
                    "max_tokens": 4000,  # 增加token数量以获得更详细的分析
                    "temperature": 0.7,  # 保持一定创造性
                    "top_p": 0.9
                }
            }
            
            print("[AI分析] 正在调用阿里千问API进行智能分析...")
            print("[AI分析] 提示: API调用可能需要30-90秒，请耐心等待...")
            
            # 增加重试机制和超时设置
            max_retries = 2
            timeout = 120  # 增加到120秒
            
            for attempt in range(max_retries):
                try:
                    if attempt > 0:
                        print(f"[AI分析] 重试第 {attempt} 次...")
                    
                    response = requests.post(api_url, headers=headers, json=payload, timeout=timeout)
                    
                    if response.status_code == 200:
                        result = response.json()
                        if 'output' in result and 'choices' in result['output']:
                            ai_analysis = result['output']['choices'][0]['message']['content']
                            print("[AI分析] [OK] AI分析完成")
                            return ai_analysis
                        else:
                            print(f"[AI分析] [WARNING] API返回格式异常: {result}")
                            if attempt < max_retries - 1:
                                continue
                            return None
                    else:
                        print(f"[AI分析] [ERROR] API调用失败: {response.status_code}")
                        if attempt < max_retries - 1:
                            continue
                        return None
                        
                except requests.exceptions.Timeout:
                    print(f"[AI分析] [WARNING] API调用超时 (尝试 {attempt + 1}/{max_retries})")
                    if attempt < max_retries - 1:
                        print("[AI分析] 正在重试...")
                        continue
                    print("[AI分析] [ERROR] 多次尝试后仍然超时，将使用备用模板")
                    return None
                    
                except requests.exceptions.RequestException as e:
                    print(f"[AI分析] [ERROR] 网络请求错误: {str(e)}")
                    if attempt < max_retries - 1:
                        continue
                    return None
            
            return None
                
        except Exception as e:
            print(f"[AI分析] [ERROR] 调用千问API失败: {str(e)}")
            import traceback
            traceback.print_exc()
            return None
    
    def _prepare_data_summary(self, user_date_range: tuple = None) -> str:
        """准备数据摘要供AI分析
        
        Args:
            user_date_range: 用户选择的时间范围 (start_date, end_date)
        """
        if self.data is None or self.data.empty:
            return "无数据"
        
        # 格式化日期 - 优先使用用户选择的时间范围
        if user_date_range:
            min_date = user_date_range[0]
            max_date = user_date_range[1]
            time_span = (pd.to_datetime(max_date) - pd.to_datetime(min_date)).days + 1
        else:
            # 如果没有用户选择，使用数据中的实际日期
            min_date = pd.to_datetime(self.data['询盘时间'].min()).strftime('%Y-%m-%d')
            max_date = pd.to_datetime(self.data['询盘时间'].max()).strftime('%Y-%m-%d')
            time_span = (pd.to_datetime(max_date) - pd.to_datetime(min_date)).days + 1
        
        total_inquiries = len(self.data)
        total_customers = self.data['客户名称'].nunique()
        total_countries = self.data['国家'].nunique()
        
        # 计算各层级询盘
        level_a_count = len(self.data[self.data['跟进等级'] == 'A'])
        level_b_count = len(self.data[self.data['跟进等级'] == 'B'])
        level_c_count = len(self.data[self.data['跟进等级'] == 'C'])
        level_x_count = len(self.data[self.data['跟进等级'] == 'X'])
        
        # 国家分布
        country_dist = self.data['国家'].value_counts().head(10)
        country_str = "\n".join([f"  {i+1}. {country}: {count}条 ({count/total_inquiries*100:.1f}%)" 
                                 for i, (country, count) in enumerate(country_dist.items())])
        
        # 产品分布
        product_dist = self.data['询价产品'].value_counts().head(10)
        product_str = "\n".join([f"  {i+1}. {product}: {count}次 ({count/total_inquiries*100:.1f}%)" 
                                 for i, (product, count) in enumerate(product_dist.items())])
        
        # 客户层级分布
        if '客户层级' in self.data.columns:
            level_dist = self.data['客户层级'].value_counts()
            level_str = "\n".join([f"  {level}: {level_dist.get(level, 0)}条 ({level_dist.get(level, 0)/total_inquiries*100:.1f}%)" 
                                   for level in ['L4', 'L3', 'L2', 'L1', 'L0']])
        else:
            level_str = "  无客户层级数据"
        
        # 咨询方式分布
        method_dist = self.data['咨询方式'].value_counts()
        method_str = "\n".join([f"  {method}: {count}条 ({count/total_inquiries*100:.1f}%)" 
                                for method, count in method_dist.items()])
        
        # 团队绩效
        handler_performance = self.data.groupby('跟进人').agg({
            '客户名称': 'count',
            '跟进等级': lambda x: (x == 'A').sum()
        }).rename(columns={'客户名称': '询盘数', '跟进等级': 'A级数'})
        handler_performance['A级占比'] = (handler_performance['A级数'] / handler_performance['询盘数'] * 100).round(1)
        handler_performance = handler_performance.sort_values('询盘数', ascending=False)
        handler_str = "\n".join([f"  {handler}: {int(row['询盘数'])}条询盘, {int(row['A级数'])}条A级 ({row['A级占比']}%)" 
                                 for handler, row in handler_performance.iterrows()])
        
        # 时间趋势
        self.data['日期'] = pd.to_datetime(self.data['询盘时间']).dt.date
        daily_trend = self.data.groupby('日期').size()
        avg_daily = daily_trend.mean()
        max_day = daily_trend.idxmax()
        max_count = daily_trend.max()
        min_day = daily_trend.idxmin()
        min_count = daily_trend.min()
        
        # 周度增长率（如果时间跨度足够）
        growth_rate_str = ""
        if time_span >= 14:
            self.data['周次'] = pd.to_datetime(self.data['询盘时间']).dt.isocalendar().week
            weekly_trend = self.data.groupby('周次').size()
            if len(weekly_trend) >= 2:
                recent_week_avg = weekly_trend.iloc[-2:].mean()
                early_week_avg = weekly_trend.iloc[:2].mean() if len(weekly_trend) > 2 else recent_week_avg
                growth_rate = ((recent_week_avg - early_week_avg) / early_week_avg * 100) if early_week_avg > 0 else 0
                growth_rate_str = f"\n• 周度增长率: {growth_rate:+.1f}% (近期周均{recent_week_avg:.1f}条 vs 早期周均{early_week_avg:.1f}条)"
        
        summary = f"""
时间范围: {min_date} 至 {max_date} (共{time_span}天)

核心指标:
• 总询盘数: {total_inquiries}条
• 日均询盘: {avg_daily:.1f}条
• 独立客户数: {total_customers}个
• 覆盖国家: {total_countries}个
• A级询盘: {level_a_count}条 ({level_a_count/total_inquiries*100:.1f}%) - 精准高价值
• B级询盘: {level_b_count}条 ({level_b_count/total_inquiries*100:.1f}%) - 有潜力
• C级询盘: {level_c_count}条 ({level_c_count/total_inquiries*100:.1f}%) - 需培育
• X级询盘: {level_x_count}条 ({level_x_count/total_inquiries*100:.1f}%) - 无效询盘

TOP 10 国家分布:
{country_str}

TOP 10 热门产品:
{product_str}

客户层级分布 (L4最高，L0最低):
{level_str}

咨询方式分布:
{method_str}

团队成员绩效:
{handler_str}

时间趋势:
• 日均询盘: {avg_daily:.1f}条
• 峰值: {max_day} ({max_count}条)
• 低谷: {min_day} ({min_count}条){growth_rate_str}
"""
        return summary
    
    def _generate_text_report(self, output_file: str, user_date_range: tuple = None):
        """生成AI驱动的专业详细文本格式报告
        
        Args:
            output_file: 输出文件路径
            user_date_range: 用户选择的时间范围 (start_date, end_date)
        """
        try:
            with open(output_file, 'w', encoding='utf-8') as f:
                f.write("=" * 100 + "\n")
                f.write("                    阿里国际业务数据分析报告\n")
                f.write("             Ali International Business Data Analysis Report\n")
                f.write("                  (Powered by 阿里千问 AI)\n")
                f.write("=" * 100 + "\n\n")
                
                if self.data is None or self.data.empty:
                    f.write("⚠️ 数据为空，无法生成报告\n")
                    return
                
                # 准备数据摘要
                print("\n" + "="*80)
                print("[报告生成] 步骤 1/3: 准备数据摘要...")
                data_summary = self._prepare_data_summary(user_date_range)
                print("[报告生成] [OK] 数据摘要准备完成")
                
                # 调用AI分析
                print("[报告生成] 步骤 2/3: 调用阿里千问AI进行智能分析...")
                ai_analysis = self._call_qianwen_api(data_summary)
                
                # 写入报告
                print("[报告生成] 步骤 3/3: 生成报告文件...")
                
                # 格式化日期 - 优先使用用户选择的时间范围
                if user_date_range:
                    min_date = user_date_range[0]
                    max_date = user_date_range[1]
                    time_span = (pd.to_datetime(max_date) - pd.to_datetime(min_date)).days + 1
                    print(f"[报告生成] 使用用户选择的时间范围: {min_date} 至 {max_date}")
                else:
                    min_date = pd.to_datetime(self.data['询盘时间'].min()).strftime('%Y-%m-%d')
                    max_date = pd.to_datetime(self.data['询盘时间'].max()).strftime('%Y-%m-%d')
                    time_span = (pd.to_datetime(max_date) - pd.to_datetime(min_date)).days + 1
                    print(f"[报告生成] 使用数据中的实际时间范围: {min_date} 至 {max_date}")
                
                f.write(f"📅 分析时段: {min_date} 至 {max_date} (共 {time_span} 天)\n")
                f.write(f"📊 报告生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"🤖 分析引擎: 阿里千问 (qwen-max)\n")
                f.write("\n" + "=" * 100 + "\n\n")
                
                if ai_analysis:
                    # 使用AI生成的分析报告
                    f.write(ai_analysis)
                    f.write("\n\n")
                    print("[报告生成] [OK] AI分析报告已写入")
                else:
                    # AI调用失败时的备用方案
                    f.write("[警告] AI分析服务暂时不可用，使用备用分析模板\n\n")
                    print("[报告生成] [WARNING] AI分析失败，使用备用模板")
                    self._generate_fallback_report(f)
                
                # 添加数据附录
                f.write("\n" + "=" * 100 + "\n")
                f.write("【数据附录 DATA APPENDIX】\n")
                f.write("=" * 100 + "\n\n")
                f.write(data_summary)
                f.write("\n")
                
                f.write("\n" + "=" * 100 + "\n")
                f.write("报告结束 END OF REPORT\n")
                f.write("=" * 100 + "\n")
                
                print("[报告生成] [OK] 报告生成完成")
                print("="*80 + "\n")
                
        except Exception as e:
            print(f"[报告生成] [ERROR] 生成文本报告失败: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def _generate_fallback_report(self, f):
        """备用报告生成方案（当AI不可用时）"""
        f.write("注意：当前使用简化模板。建议配置阿里千问API以获得更智能的分析。\n\n")
        f.write("【简要分析】\n\n")
        
        total_inquiries = len(self.data)
        level_a_count = len(self.data[self.data['跟进等级'] == 'A'])
        
        f.write(f"本期共收到 {total_inquiries} 条询盘，其中A级高价值询盘 {level_a_count} 条。\n")
        f.write("建议重点关注A级和B级客户，优化产品推广策略，加强市场拓展。\n\n")
        f.write("详细数据请查看下方数据附录。\n")
    
    def _analyze_high_value_customers(self):
        """分析高价值客户"""
        # 这个方法已被AI分析替代，保留以避免破坏其他代码
        return []
    
    def _analyze_priority_customers(self):
        """分析需要重点跟进的客户"""
        # 这个方法已被AI分析替代，保留以避免破坏其他代码
        return []
    
    def _analyze_lost_customers(self):
        """分析失单原因"""
        # 这个方法已被AI分析替代，保留以避免破坏其他代码
        return {}
    
    def _generate_recommendations(self):
        """生成发展建议"""
        # 这个方法已被AI分析替代，保留以避免破坏其他代码
        return []
    
    # 原有的辅助方法已被移除，新报告完全由AI生成
    
    # =========================
    # 智能提醒
    # =========================
    
    def get_alerts(self) -> List[Dict]:
        """生成智能提醒"""
        alerts = []
        
        if self.data is None:
            return alerts
        
        current_date = datetime.now()
        
        # 检查长时间未跟进的客户
        if '最后跟进时间' in self.data.columns:
            overdue_customers = self.data[
                pd.notna(self.data['最后跟进时间'])
            ]
            
            for _, customer in overdue_customers.iterrows():
                last_follow_up = pd.to_datetime(customer['最后跟进时间'])
                days_since = (current_date - last_follow_up).days
                
                if days_since > 7:  # 超过7天未跟进
                    alerts.append({
                        'type': 'follow_up_overdue',
                        'priority': 'high' if days_since > 14 else 'medium',
                        'message': f"客户 {customer.get('客户名称', '未知')} 已 {days_since} 天未跟进",
                        'customer_name': customer.get('客户名称', '未知'),
                        'days_overdue': days_since
                    })
        
        # 检查高价值客户
        if '跟进等级' in self.data.columns:
            high_value_customers = self.data[
                self.data['跟进等级'].isin(['A'])
            ]
            
            for _, customer in high_value_customers.iterrows():
                alerts.append({
                    'type': 'high_value_customer',
                    'priority': 'medium',
                    'message': f"高价值客户 {customer.get('客户名称', '未知')} 需要特别关注",
                    'customer_name': customer.get('客户名称', '未知'),
                    'level': customer.get('跟进等级'),
                    'continent': customer.get('所属大洲', '未知')
                })
        
        return alerts[:10]  # 只返回前10个提醒
    def _analyze_high_value_customers(self):
        """分析高价值客户"""
        high_value = []
        if self.data is not None:
            # 按国家分组，找出询盘数量多的国家
            country_counts = self.data['国家'].value_counts()
            top_countries = country_counts.head(5)
            
            for country, count in top_countries.items():
                country_data = self.data[self.data['国家'] == country]
                # 找出该国家的主要客户
                customer_counts = country_data['客户名称'].value_counts()
                top_customer = customer_counts.index[0] if len(customer_counts) > 0 else "未知"
                high_value.append({
                    'name': top_customer,
                    'country': country,
                    'reason': f"来自主要市场，询盘量 {count} 次 (From major market, {count} inquiries)"
                })
        
        return high_value
    
    def _analyze_priority_customers(self):
        """分析需要重点跟进的客户"""
        priority = []
        if self.data is not None:
            # 找出跟进等级为A的客户
            level_a_customers = self.data[self.data['跟进等级'] == 'A']
            if not level_a_customers.empty:
                for _, customer in level_a_customers.iterrows():
                    priority.append({
                        'name': customer['客户名称'],
                        'country': customer['国家'],
                        'reason': "精准询盘，高转化潜力 (Precise inquiry, high conversion potential)"
                    })
            
            # 找出最近询盘但未跟进的客户
            recent_date = pd.Timestamp(datetime.now().date()) - pd.Timedelta(days=7)
            recent_customers = self.data[self.data['询盘时间'] >= recent_date]
            for _, customer in recent_customers.iterrows():
                if customer['客户名称'] not in [p['name'] for p in priority]:
                    priority.append({
                        'name': customer['客户名称'],
                        'country': customer['国家'],
                        'reason': "最近询盘，需要及时跟进 (Recent inquiry, timely follow-up needed)"
                    })
        
        return priority[:10]  # 限制数量
    
    def _analyze_lost_customers(self):
        """分析失单原因"""
        lost_reasons = {}
        if self.data is not None and '备注 (失单原因+跟进机会点)' in self.data.columns:
            remarks = self.data['备注 (失单原因+跟进机会点)'].dropna()
            for remark in remarks:
                if isinstance(remark, str) and remark.strip():
                    # 简单的关键词分析
                    if '不回' in remark or '未读' in remark:
                        lost_reasons['客户不回复'] = lost_reasons.get('客户不回复', 0) + 1
                    elif '价格' in remark:
                        lost_reasons['价格问题'] = lost_reasons.get('价格问题', 0) + 1
                    elif 'MOQ' in remark or 'moq' in remark:
                        lost_reasons['起订量问题'] = lost_reasons.get('起订量问题', 0) + 1
                    elif '个人' in remark:
                        lost_reasons['个人买家'] = lost_reasons.get('个人买家', 0) + 1
                    else:
                        lost_reasons['其他原因'] = lost_reasons.get('其他原因', 0) + 1
        
        return lost_reasons
    
    def _generate_recommendations(self):
        """生成发展建议"""
        recommendations = [
            "加强高价值市场的客户维护，建立长期合作关系 (Strengthen customer maintenance in high-value markets and establish long-term partnerships)",
            "优化价格策略，提高竞争力 (Optimize pricing strategy to improve competitiveness)",
            "降低起订量要求，吸引更多中小客户 (Reduce MOQ requirements to attract more small and medium-sized customers)",
            "建立客户分级管理体系，差异化服务 (Establish customer tiering management system with differentiated services)",
            "加强产品宣传，提高品牌知名度 (Strengthen product promotion to enhance brand awareness)",
            "建立客户反馈机制，持续改进服务质量 (Establish customer feedback mechanism to continuously improve service quality)",
            "开发新产品线，扩大市场覆盖 (Develop new product lines to expand market coverage)",
            "加强团队培训，提高跟进效率 (Strengthen team training to improve follow-up efficiency)"
        ]
        return recommendations
    
    def get_smart_alerts(self) -> List[Dict]:
        """获取智能提醒 - 6大核心提醒类别"""
        if self.data is None:
            raise ValueError("请先读取数据")
        
        alerts = []
        current_date = datetime.now()
        
        # 第一类：高价值客户识别提醒（抓住黄金机会）
        alerts.extend(self._check_high_value_customers())
        
        # 第二类：垃圾/钓鱼/低质量询盘预警
        alerts.extend(self._check_low_quality_inquiries())
        
        # 第三类：长期未跟进提醒（防止客户流失）
        alerts.extend(self._check_long_term_unfollow(current_date))
        
        # 第四类：区域集中趋势提醒（把握市场动向）
        alerts.extend(self._check_regional_trends())
        
        # 第五类：产品热度变化提醒（指导备货与推广）
        alerts.extend(self._check_product_trends())
        
        # 第六类：转化漏斗异常提醒（优化销售策略）
        alerts.extend(self._check_conversion_funnel())
        
        # 按优先级排序
        priority_order = {'high': 0, 'medium': 1, 'low': 2}
        alerts.sort(key=lambda x: priority_order.get(x['priority'], 3))
        
        return alerts
    
    def _check_high_value_customers(self) -> List[Dict]:
        """第一类：高价值客户识别提醒"""
        alerts = []
        
        if '备注 (失单原因+跟进机会点)' not in self.data.columns:
            return alerts
        
        # 高价值关键词
        high_value_keywords = [
            '自有设计图', '定制品牌', '首单100件', '首单80件', '首单50件',
            '官网', '线上店铺', 'OEM', '品牌定制', '大单', '长期合作',
            'wholesale', 'bulk order', 'brand', 'custom', 'private label'
        ]
        
        for idx, row in self.data.iterrows():
            remark = str(row.get('备注 (失单原因+跟进机会点)', '')).lower()
            customer_name = row.get('客户名称', '未知')
            country = row.get('国家', '未知')
            
            for keyword in high_value_keywords:
                if keyword.lower() in remark:
                    alerts.append({
                        'type': 'high_value_opportunity',
                        'priority': 'high',
                        'category': '🚨 高价值客户识别',
                        'message': f"[高潜力客户] {customer_name} ({country}) - 备注提及'{keyword}'",
                        'suggestion': '立即电话或TM联系，提供OEM报价模板，发送成功案例',
                        'customer_name': customer_name,
                        'country': country,
                        'keyword': keyword
                    })
                    break
        
        # 检查跟进等级升级（B/C 升为 X）
        if '跟进等级' in self.data.columns:
            x_level_customers = self.data[self.data['跟进等级'] == 'X']
            for _, customer in x_level_customers.head(5).iterrows():
                alerts.append({
                    'type': 'level_upgraded',
                    'priority': 'high',
                    'category': '🚨 高价值客户识别',
                    'message': f"[升级提醒] {customer.get('客户名称', '未知')} ({customer.get('国家', '未知')}) 已进入样品阶段",
                    'suggestion': '创建任务：寄样+发票+物流单号录入，设置7天后回访',
                    'customer_name': customer.get('客户名称', '未知'),
                    'level': 'X'
                })
        
        return alerts
    
    def _check_low_quality_inquiries(self) -> List[Dict]:
        """第二类：垃圾/钓鱼/低质量询盘预警"""
        alerts = []
        
        if '备注 (失单原因+跟进机会点)' not in self.data.columns:
            return alerts
        
        # 低质量关键词
        low_quality_keywords = [
            '钓鱼', '新注册用户未读', '促销商', '一句话询盘', '不对口',
            '个人买家', '垃圾询盘', '无效询盘', '诈骗', '骗子'
        ]
        
        low_quality_count = 0
        for idx, row in self.data.iterrows():
            remark = str(row.get('备注 (失单原因+跟进机会点)', '')).lower()
            customer_name = row.get('客户名称', '未知')
            country = row.get('国家', '未知')
            level = row.get('客户层级', 'L0')
            
            for keyword in low_quality_keywords:
                if keyword.lower() in remark:
                    low_quality_count += 1
                    alerts.append({
                        'type': 'low_quality_warning',
                        'priority': 'low',
                        'category': '🛑 低质量询盘预警',
                        'message': f"[低质预警] {customer_name} ({country}) - 标记为\"{keyword}\"",
                        'suggestion': '标记为C级，归入观察池，不投入深度沟通资源',
                        'customer_name': customer_name,
                        'country': country
                    })
                    break
        
        # 添加汇总提醒
        if low_quality_count > 0:
            c_level_count = len(self.data[self.data['跟进等级'] == 'C']) if '跟进等级' in self.data.columns else 0
            total_count = len(self.data)
            c_level_percentage = (c_level_count / total_count * 100) if total_count > 0 else 0
            
            alerts.append({
                'type': 'low_quality_summary',
                'priority': 'medium',
                'category': '🛑 低质量询盘预警',
                'message': f"[数据洞察] 当前C级占比{c_level_percentage:.1f}%，智能过滤可节省约{c_level_percentage * 0.6:.0f}%无效沟通时间",
                'suggestion': f'发现{low_quality_count}个低质量询盘，建议优化筛选策略',
                'count': low_quality_count
            })
        
        return alerts
    
    def _check_long_term_unfollow(self, current_date) -> List[Dict]:
        """第三类：长期未跟进提醒（防止客户流失）"""
        alerts = []
        
        if '最后跟进时间' not in self.data.columns:
            return alerts
        
        # 检查超过5天未跟进且非X/A级的客户
        for idx, row in self.data.iterrows():
            last_followup = row.get('最后跟进时间')
            level = row.get('跟进等级', '')
            customer_name = row.get('客户名称', '未知')
            country = row.get('国家', '未知')
            
            if pd.notna(last_followup) and level not in ['X', 'A']:
                try:
                    last_date = pd.to_datetime(last_followup)
                    days_overdue = (current_date - last_date).days
                    
                    if days_overdue > 5:
                        alerts.append({
                            'type': 'long_term_unfollow',
                            'priority': 'high' if days_overdue > 7 else 'medium',
                            'category': '🔁 长期未跟进提醒',
                            'message': f"[滞留提醒] {customer_name} ({country}) 已{days_overdue}天未回复",
                            'suggestion': '重新触达，尝试换主题邮件或TM消息: "上次提到的价格是否合适？我们可以调整MOQ方案。"',
                            'customer_name': customer_name,
                            'days_overdue': days_overdue
                        })
                except:
                    pass
        
        # 检查客户曾表达兴趣但"未读"消息
        if '备注 (失单原因+跟进机会点)' in self.data.columns:
            for idx, row in self.data.iterrows():
                remark = str(row.get('备注 (失单原因+跟进机会点)', '')).lower()
                if '未读' in remark and ('样品' in remark or '兴趣' in remark):
                    customer_name = row.get('客户名称', '未知')
                    country = row.get('国家', '未知')
                    alerts.append({
                        'type': 'unread_message',
                        'priority': 'medium',
                        'category': '🔁 长期未跟进提醒',
                        'message': f"[唤醒提醒] {customer_name} ({country}) 曾寻求样品但未读信息",
                        'suggestion': '使用"免样品费门槛"作为钩子重新激活',
                        'customer_name': customer_name
                    })
        
        return alerts
    
    def _check_regional_trends(self) -> List[Dict]:
        """第四类：区域集中趋势提醒（把握市场动向）"""
        alerts = []
        
        if '国家' not in self.data.columns or '询盘时间' not in self.data.columns:
            return alerts
        
        # 检查近两周的询盘
        two_weeks_ago = datetime.now() - timedelta(days=14)
        recent_data = self.data[pd.to_datetime(self.data['询盘时间']) >= two_weeks_ago]
        
        # 统计国家频次
        country_counts = recent_data['国家'].value_counts()
        
        # 检查同一国家连续出现3次及以上
        for country, count in country_counts.items():
            if count >= 3:
                # 查找该国家的热门产品
                country_data = recent_data[recent_data['国家'] == country]
                if '询价产品' in country_data.columns:
                    product_counts = country_data['询价产品'].value_counts()
                    top_product = product_counts.index[0] if len(product_counts) > 0 else '未知'
                    
                    alerts.append({
                        'type': 'regional_hotspot',
                        'priority': 'high',
                        'category': '🌍 区域集中趋势',
                        'message': f"[区域热点] 近两周{country}出现{count}次询盘，集中在{top_product}",
                        'suggestion': f'准备{country}本地化文案；检查库存与物流方案',
                        'country': country,
                        'count': count,
                        'product': top_product
                    })
        
        # 检查大洲客户咨询量周环比增长
        if '所属大洲' in self.data.columns:
            one_week_ago = datetime.now() - timedelta(days=7)
            this_week_data = self.data[pd.to_datetime(self.data['询盘时间']) >= one_week_ago]
            last_week_data = self.data[
                (pd.to_datetime(self.data['询盘时间']) >= two_weeks_ago) &
                (pd.to_datetime(self.data['询盘时间']) < one_week_ago)
            ]
            
            this_week_continent = this_week_data['所属大洲'].value_counts()
            last_week_continent = last_week_data['所属大洲'].value_counts()
            
            for continent in this_week_continent.index:
                this_count = this_week_continent.get(continent, 0)
                last_count = last_week_continent.get(continent, 0)
                
                if last_count > 0:
                    growth_rate = ((this_count - last_count) / last_count) * 100
                    if growth_rate > 50:
                        alerts.append({
                            'type': 'emerging_market',
                            'priority': 'high',
                            'category': '🌍 区域集中趋势',
                            'message': f"[新兴市场] {continent}客户数量本周上升{growth_rate:.0f}%",
                            'suggestion': f'优化{continent}物流方案，更新运费计算器',
                            'continent': continent,
                            'growth_rate': growth_rate
                        })
        
        return alerts
    
    def _check_product_trends(self) -> List[Dict]:
        """第五类：产品热度变化提醒（指导备货与推广）"""
        alerts = []
        
        if '询价产品' not in self.data.columns or '询盘时间' not in self.data.columns:
            return alerts
        
        # 检查近一周的产品询价
        one_week_ago = datetime.now() - timedelta(days=7)
        recent_data = self.data[pd.to_datetime(self.data['询盘时间']) >= one_week_ago]
        
        product_counts = recent_data['询价产品'].value_counts()
        
        # 检查某产品被提及 ≥3次/周
        for product, count in product_counts.items():
            if count >= 3:
                # 统计询问该产品的国家数
                product_data = recent_data[recent_data['询价产品'] == product]
                country_count = product_data['国家'].nunique()
                
                alerts.append({
                    'type': 'hot_product',
                    'priority': 'high',
                    'category': '🧩 产品热度变化',
                    'message': f"[爆款预警] {product} 近期被{country_count}个国家客户询问{count}次",
                    'suggestion': '确保该款打样资料齐全、MOQ灵活；主推此款做专题页',
                    'product': product,
                    'count': count,
                    'country_count': country_count
                })
        
        # 检查多个客户提及"低MOQ"需求
        if '备注 (失单原因+跟进机会点)' in self.data.columns:
            low_moq_count = 0
            for idx, row in recent_data.iterrows():
                remark = str(row.get('备注 (失单原因+跟进机会点)', '')).lower()
                if 'moq' in remark or '起订量' in remark or '小批量' in remark:
                    low_moq_count += 1
            
            if low_moq_count >= 5:
                alerts.append({
                    'type': 'low_moq_demand',
                    'priority': 'medium',
                    'category': '🧩 产品热度变化',
                    'message': f"[需求洞察] 本周有{low_moq_count}位客户明确表示'低MOQ'需求",
                    'suggestion': '推出"Mini MOQ Package"服务（如50件起订），差异化竞争',
                    'count': low_moq_count
                })
        
        return alerts
    
    def _check_conversion_funnel(self) -> List[Dict]:
        """第六类：转化漏斗异常提醒（优化销售策略）"""
        alerts = []
        
        if '跟进等级' not in self.data.columns or '询盘时间' not in self.data.columns:
            return alerts
        
        # 检查近两周的数据
        two_weeks_ago = datetime.now() - timedelta(days=14)
        recent_data = self.data[pd.to_datetime(self.data['询盘时间']) >= two_weeks_ago]
        
        if len(recent_data) == 0:
            return alerts
        
        # 检查A级精准询盘数量
        a_level_count = len(recent_data[recent_data['跟进等级'] == 'A'])
        
        if a_level_count == 0:
            alerts.append({
                'type': 'no_a_level',
                'priority': 'high',
                'category': '📉 转化漏斗异常',
                'message': '[漏斗警报] 过去14天无A级精准询盘，源头质量可能下降',
                'suggestion': '复盘RFQ标题与产品描述，增加"Wholesale"、"Bulk Order"等关键词',
                'count': 0
            })
        
        # 检查X级客户比例
        x_level_count = len(recent_data[recent_data['跟进等级'] == 'X'])
        x_level_percentage = (x_level_count / len(recent_data)) * 100
        
        if x_level_percentage < 5:
            alerts.append({
                'type': 'low_conversion',
                'priority': 'high',
                'category': '📉 转化漏斗异常',
                'message': f'[转化瓶颈] 仅{x_level_percentage:.1f}%客户进入样品阶段，转化率偏低',
                'suggestion': '增加客户见证视频、第三方检测报告、工厂实拍增强可信度',
                'percentage': x_level_percentage
            })
        
        # 检查C级占比
        c_level_count = len(recent_data[recent_data['跟进等级'] == 'C'])
        c_level_percentage = (c_level_count / len(recent_data)) * 100
        
        if c_level_percentage > 50:
            alerts.append({
                'type': 'high_c_level',
                'priority': 'medium',
                'category': '📉 转化漏斗异常',
                'message': f'[质量预警] C级询盘占比{c_level_percentage:.0f}%，源头筛选需优化',
                'suggestion': '优化RFQ自动回复规则，提高初筛门槛',
                'percentage': c_level_percentage
            })
        
        return alerts
    
    def export_data(self, output_file: str, format: str = 'excel', group_by_month: bool = True):
        """导出数据 - 完全按照用户提供的模板格式，按月分组"""
        if self.data is None:
            raise ValueError("请先读取数据")
        
        os.makedirs(os.path.dirname(output_file) if os.path.dirname(output_file) else '.', exist_ok=True)
        
        if format.lower() == 'excel':
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            wb = Workbook()
            # 删除默认工作表
            wb.remove(wb.active)
            
            # 定义列名和对应的颜色
            columns_info = {
                '询盘时间': '0000FF',      # Blue
                '咨询方式': '0000FF',      # Blue
                '跟进等级': 'FF0000',      # Red
                '客户名称': 'FFFF00',      # Yellow
                '客户层级': 'FFFF00',      # Yellow
                '所属大洲': 'FF0000',      # Red
                '国家': 'FF0000',          # Red
                '询价产品': 'FF0000',      # Red
                '产品ID': 'FF0000',        # Red
                '跟进人': '0000FF',        # Blue
                '备注 (失单原因+跟进机会点)': '0000FF',  # Blue
                '最后跟进时间': '0000FF'    # Blue
            }
            
            # 设置边框样式
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            if group_by_month and '询盘时间' in self.data.columns:
                # 按月份分组
                def extract_month(date_str):
                    try:
                        date_obj = pd.to_datetime(date_str, errors='coerce')
                        if pd.notna(date_obj):
                            return date_obj.strftime('%Y年%m月')
                        return '未知'
                    except:
                        return '未知'
                
                self.data['_月份'] = self.data['询盘时间'].apply(extract_month)
                month_groups = self.data.groupby('_月份')
                
                for month, month_data in month_groups:
                    # 创建月度工作表
                    ws = wb.create_sheet(title=month)
                    
                    # 写入表头
                    self._write_header(ws, columns_info, thin_border)
                    
                    # 写入数据
                    self._write_data(ws, month_data, thin_border)
                    
                    # 添加自动筛选和冻结窗格
                    self._apply_excel_features(ws, len(month_data))
                    
                    # 添加分类规则
                    self._add_classification_rules(ws, len(self.standard_columns))
                    
                    self.logger.info(f"已创建工作表 '{month}'，包含 {len(month_data)} 条记录")
                    
                # 移除辅助列
                self.data = self.data.drop(columns=['_月份'])
            else:
                # 不分组，所有数据在一个工作表
                ws = wb.create_sheet(title="客户跟进表")
                
                # 写入表头
                self._write_header(ws, columns_info, thin_border)
                
                # 写入数据
                self._write_data(ws, self.data, thin_border)
                
                # 添加自动筛选和冻结窗格
                self._apply_excel_features(ws, len(self.data))
                
                # 添加分类规则
                self._write_classification_rules(ws, len(self.standard_columns))
            
            # 添加图表工作表
            self._add_charts_to_excel(wb, output_file)
            
            # 保存文件
            wb.save(output_file)
            
        elif format.lower() == 'csv':
            self.data.to_csv(output_file, index=False, encoding='utf-8-sig')
        else:
            raise ValueError("不支持的导出格式")
        
        self.logger.info(f"数据已导出到: {output_file}")
    
    def _write_header(self, ws, columns_info, thin_border):
        """写入表头并设置样式"""
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import PatternFill, Font, Alignment, Border
        
        for col_idx, col_name in enumerate(self.standard_columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            
            # 设置背景颜色
            hex_color = columns_info.get(col_name, 'FFFFFF')
            cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
            
            # 设置字体：白色、加粗
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            
            # 居中
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 设置边框
            cell.border = thin_border
            
            # 设置列宽（根据内容调整）
            if col_name == '备注 (失单原因+跟进机会点)':
                ws.column_dimensions[get_column_letter(col_idx)].width = 40
            elif col_name in ['客户名称', '询价产品']:
                ws.column_dimensions[get_column_letter(col_idx)].width = 20
            elif col_name in ['询盘时间', '最后跟进时间']:
                ws.column_dimensions[get_column_letter(col_idx)].width = 12
            else:
                ws.column_dimensions[get_column_letter(col_idx)].width = 15
        
        # 设置表头行高
        ws.row_dimensions[1].height = 25
    
    def _apply_excel_features(self, ws, data_rows):
        """应用Excel特性：自动筛选、冻结窗格、隔行填充"""
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import PatternFill
        
        # 1. 添加自动筛选（从A1到最后一列的最后一行）
        last_col = get_column_letter(len(self.standard_columns))
        last_row = data_rows + 1  # 表头占1行
        ws.auto_filter.ref = f'A1:{last_col}{last_row}'
        
        # 2. 冻结首行（表头）
        ws.freeze_panes = 'A2'
        
        # 3. 隔行填充浅灰色（提高可读性）
        light_gray_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        for row_idx in range(3, last_row + 1, 2):  # 从第3行开始，每隔一行
            for col_idx in range(1, len(self.standard_columns) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                # 只在没有特殊背景色的单元格上应用
                if not cell.fill or cell.fill.start_color.rgb == '00000000':
                    cell.fill = light_gray_fill
        
        # 4. 设置打印选项（可选）
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE  # 横向
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToHeight = False
        ws.page_setup.fitToWidth = 1
        
        # 5. 设置缩放（可选，默认100%）
        ws.sheet_view.zoomScale = 100
        
        self.logger.info(f"已应用Excel特性：自动筛选范围 A1:{last_col}{last_row}，冻结首行")
    
    def _write_data(self, ws, data, thin_border):
        """写入数据行"""
        from openpyxl.styles import Font
        from openpyxl.cell.cell import TYPE_STRING
        from openpyxl.utils import get_column_letter
        
        for row_idx, (_, row_data) in enumerate(data.iterrows(), 2):
            for col_idx, col_name in enumerate(self.standard_columns, 1):
                cell_value = row_data.get(col_name, '')
                
                # 特殊处理询盘时间，只显示日期不显示时间
                if col_name == '询盘时间':
                    if pd.notna(cell_value) and cell_value != '':
                        try:
                            if isinstance(cell_value, pd.Timestamp):
                                cell_value = cell_value.strftime('%Y-%m-%d')
                            elif isinstance(cell_value, str):
                                # 尝试解析字符串并只取日期部分
                                date_obj = pd.to_datetime(cell_value, errors='coerce')
                                if pd.notna(date_obj):
                                    cell_value = date_obj.strftime('%Y-%m-%d')
                        except Exception:
                            pass
                
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                
                # 设置边框
                cell.border = thin_border
                
                # 设置对齐方式
                from openpyxl.styles import Alignment
                if col_name == '备注 (失单原因+跟进机会点)':
                    # 备注列：左对齐，自动换行
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                elif col_name in ['询盘时间', '最后跟进时间', '跟进等级', '客户层级', '咨询方式']:
                    # 时间和等级列：居中对齐
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    # 其他列：左对齐
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                
                # 产品ID必须设为文本格式（防止显示为科学计数法）
                if col_name == '产品ID':
                    cell.data_type = TYPE_STRING
                    # 如果是数字，转换为字符串
                    if isinstance(cell_value, (int, float)) and not pd.isna(cell_value):
                        cell.value = str(int(cell_value))  # 转换为整数再转字符串，保持原始格式
                
                # 最后跟进时间保持原始格式（只显示日期）
                if col_name == '最后跟进时间':
                    # 如果是datetime类型，转换为字符串（只显示日期）
                    if pd.notna(cell_value) and cell_value != '':
                        try:
                            if isinstance(cell_value, pd.Timestamp):
                                cell.value = cell_value.strftime('%Y-%m-%d')
                            elif isinstance(cell_value, str):
                                # 尝试解析字符串并只取日期部分
                                date_obj = pd.to_datetime(cell_value, errors='coerce')
                                if pd.notna(date_obj):
                                    cell.value = date_obj.strftime('%Y-%m-%d')
                                else:
                                    cell.value = cell_value
                        except Exception:
                            pass
                
                # 如果是指定列，设置红色字体
                if col_name in ['跟进等级', '所属大洲', '国家', '询价产品', '产品ID']:
                    cell.font = Font(color="FF0000")
    
    def _add_classification_rules(self, ws, standard_col_count):
        """添加分类规则说明"""
        self._write_classification_rules(ws, standard_col_count)
    
    def _write_classification_rules(self, ws, standard_col_count):
        """写入分类规则说明"""
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter
        
        # 分类规则
        classification_rules = [
            ("A: 精准询盘", "客户明确指出产品需求, 包含各种信息 (数量、运输/支付要求、公司信息等)"),
            ("B: 普通询盘", "广撒网询盘, 内容广泛, 只是询价或发对产品感兴趣, 或信息未读, 需要继续跟进了"),
            ("C: 个人买家/不匹配询盘/垃圾询盘", ""),
            ("X: 已下样品单/大货客户, 持续跟进", "")
        ]
        
        start_col_rules = standard_col_count + 2
        
        # 写入规则标题
        ws.cell(row=1, column=start_col_rules, value="跟进等级分类说明").font = Font(bold=True)
        ws.cell(row=1, column=start_col_rules).alignment = Alignment(horizontal='center')
        
        # 写入规则内容
        for idx, (level, desc) in enumerate(classification_rules, 2):
            ws.cell(row=idx, column=start_col_rules, value=level).font = Font(bold=True)
            if desc:
                ws.cell(row=idx, column=start_col_rules + 1, value=desc)
        
        # 调整规则说明的列宽
        ws.column_dimensions[get_column_letter(start_col_rules)].width = 25
        ws.column_dimensions[get_column_letter(start_col_rules + 1)].width = 50
    
    def _add_charts_to_excel(self, wb, output_file):
        """添加图表到Excel工作表"""
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.styles import Font
        import os
        
        # 先生成所有图表到output目录
        chart_dir = './output'
        try:
            # 生成图表（如果还没有生成）
            if not os.path.exists(os.path.join(chart_dir, 'country_dist.png')):
                self.logger.info("正在生成图表...")
                self.generate_visualizations(chart_dir)
            
            # 创建图表工作表
            ws_charts = wb.create_sheet(title="📊数据图表")
            
            # 图表文件列表（对应generate_visualizations生成的文件名）
            charts_info = [
                {'file': 'country_distribution.png', 'title': '国家分布TOP15'},
                {'file': 'follow_up_distribution.png', 'title': '跟进等级分布'},
                {'file': 'product_popularity.png', 'title': '产品分布TOP10'},
                {'file': 'daily_trend.png', 'title': '每日询盘趋势'},
                {'file': 'continent_distribution.png', 'title': '大洲分布'},
            ]
            
            current_row = 2
            for chart_info in charts_info:
                chart_file = os.path.join(chart_dir, chart_info['file'])
                
                if os.path.exists(chart_file):
                    # 添加标题
                    title_cell = ws_charts.cell(row=current_row, column=2)
                    title_cell.value = chart_info['title']
                    title_cell.font = Font(name='微软雅黑', size=14, bold=True, color='0066CC')
                    
                    # 插入图片
                    try:
                        img = XLImage(chart_file)
                        # 调整图片大小
                        img.width = 600
                        img.height = 400
                        
                        # 计算插入位置（标题下方）
                        img_position = f'B{current_row + 1}'
                        ws_charts.add_image(img, img_position)
                        
                        self.logger.info(f"[OK] 图表 '{chart_info['title']}' 已添加到Excel")
                        
                        # 更新行位置（图片高度约20行 + 3行间距）
                        current_row += 23
                        
                    except Exception as e:
                        self.logger.warning(f"无法添加图表 {chart_file}: {e}")
                        current_row += 3
                else:
                    self.logger.warning(f"图表文件不存在: {chart_file}")
            
            # 调整列宽
            ws_charts.column_dimensions['B'].width = 80
            
            self.logger.info("[OK] 所有图表已添加到Excel的'📊数据图表'工作表")
            
        except Exception as e:
            self.logger.error(f"添加图表到Excel时出错: {e}")
            import traceback
            traceback.print_exc()
            # 不影响主要的数据导出功能


def main():
    """命令行接口"""
    parser = argparse.ArgumentParser(description='阿里国际业务智能复盘工具')
    parser.add_argument('--import', dest='import_file', help='导入Excel文件')
    parser.add_argument('--export', dest='export_file', help='导出文件路径')
    parser.add_argument('--analyze', action='store_true', help='执行数据分析')
    parser.add_argument('--report', dest='report_file', help='生成报告文件')
    parser.add_argument('--visualize', action='store_true', help='生成可视化图表')
    parser.add_argument('--alerts', action='store_true', help='显示智能提醒')
    parser.add_argument('--config', default='config.json', help='配置文件路径')
    parser.add_argument('--gui', action='store_true', help='启动GUI界面')
    
    args = parser.parse_args()
    
    # 创建分析器实例
    analyzer = AliBusinessAnalyzer(args.config)
    
    if args.gui:
        # 启动GUI界面
        from ali_business_gui import AliBusinessGUI
        app = AliBusinessGUI(analyzer)
        app.run()
    else:
        # 命令行模式
        try:
            if args.import_file:
                analyzer.read_excel(args.import_file)
                print(f"成功导入数据: {len(analyzer.data)} 条记录")
            
            if args.analyze:
                results = analyzer.analyze_data()
                print("数据分析完成")
                print(f"总客户数: {results['basic_stats']['total_customers']}")
            
            if args.visualize:
                analyzer.generate_visualizations()
                print("可视化图表已生成")
            
            if args.report_file:
                report_file = analyzer.generate_report(args.report_file)
                print(f"报告已生成: {report_file}")
            
            if args.alerts:
                alerts = analyzer.get_smart_alerts()
                print(f"发现 {len(alerts)} 个提醒:")
                for alert in alerts:
                    print(f"- {alert['message']}")
            
            if args.export_file:
                analyzer.export_data(args.export_file)
                print(f"数据已导出: {args.export_file}")
                
        except Exception as e:
            print(f"执行失败: {e}")
            analyzer.logger.error(f"执行失败: {e}")


if __name__ == "__main__":
    main()
